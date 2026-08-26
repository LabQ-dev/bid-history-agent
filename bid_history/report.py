"""검색 결과를 콘솔 표 / CSV / XLSX로 출력."""

import csv
from pathlib import Path

COLUMNS = [
    ("opengDt",        "개찰일시"),
    ("_bizType",       "업무구분"),
    ("bidNtceNo",      "공고번호"),
    ("bidNtceNm",      "공고명"),
    ("dminsttNm",      "수요기관"),
    ("prtcptCnum",     "참가업체수"),
    ("opengRank",      "평가순위"),
    ("prcbdrNm",       "업체명"),
    ("bidprcAmt",      "투찰금액"),
    ("bidprcrt",       "투찰률(%)"),
    ("techEvlVal",     "기술점수"),
    ("bidPrceEvlVal",  "가격점수"),
    ("totalEvlAmtVal", "종합점수"),
    # 1위(낙찰자) 비교 — _winner에서 계산
    ("_winnerNm",      "1위업체"),
    ("_winnerAmt",     "1위투찰금액"),
    ("_winnerTotal",   "1위종합점수"),
    ("_amtDiff",       "금액차(1위대비)"),
    ("_scoreDiff",     "점수차(1위대비)"),
]


def _num(v):
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return None


def _fmt_amt(v):
    n = _num(v)
    return f"{n:,.0f}" if n is not None else ""


def enrich(r: dict) -> dict:
    """1위(낙찰자) 비교 필드를 계산해 붙인다."""
    w = r.get("_winner") or {}
    is_self_winner = str(r.get("opengRank")) == "1"
    out = dict(r)
    out["_winnerNm"] = w.get("prcbdrNm", "")
    out["_winnerAmt"] = _fmt_amt(w.get("bidprcAmt"))
    out["_winnerTotal"] = w.get("totalEvlAmtVal", "")
    out["bidprcAmt"] = _fmt_amt(r.get("bidprcAmt"))
    if is_self_winner:
        out["_amtDiff"] = "-"
        out["_scoreDiff"] = "-"
        return out
    my_amt, w_amt = _num(r.get("bidprcAmt")), _num(w.get("bidprcAmt"))
    my_sc, w_sc = _num(r.get("totalEvlAmtVal")), _num(w.get("totalEvlAmtVal"))
    out["_amtDiff"] = f"{my_amt - w_amt:+,.0f}" if (my_amt is not None and w_amt is not None) else ""
    out["_scoreDiff"] = f"{my_sc - w_sc:+.4g}" if (my_sc is not None and w_sc is not None) else ""
    return out


def to_rows(results: list[dict]) -> list[list]:
    rows = [[label for _, label in COLUMNS]]
    for r in results:
        e = enrich(r)
        rows.append([e.get(key, "") for key, _ in COLUMNS])
    return rows


def print_table(results: list[dict], limit: int = 50):
    if not results:
        print("조건에 맞는 입찰참가 이력이 없습니다.")
        return
    for raw in results[:limit]:
        r = enrich(raw)
        win = " ★낙찰(1위)" if str(r.get("opengRank")) == "1" else ""
        print(f"[{r.get('opengDt','')}] {r.get('_bizType','')} {r.get('bidNtceNo','')} "
              f"{str(r.get('bidNtceNm',''))[:40]}")
        print(f"    순위 {r.get('opengRank','-')}/{r.get('prtcptCnum','-')}{win} | "
              f"{r.get('prcbdrNm','')} | 투찰 {r.get('bidprcAmt','')}원 ({r.get('bidprcrt','')}%) | "
              f"기술 {r.get('techEvlVal','') or '-'} / 가격 {r.get('bidPrceEvlVal','') or '-'} "
              f"/ 종합 {r.get('totalEvlAmtVal','') or '-'}")
        if win == "" and r.get("_winnerNm"):
            print(f"    1위: {r['_winnerNm']} | 투찰 {r['_winnerAmt']}원 | 종합 {r['_winnerTotal'] or '-'} "
                  f"| 금액차 {r['_amtDiff'] or '-'} / 점수차 {r['_scoreDiff'] or '-'}")
    if len(results) > limit:
        print(f"... 외 {len(results) - limit}건 (파일 출력 참고)")


def save_csv(results: list[dict], path: Path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f).writerows(to_rows(results))


def save_xlsx(results: list[dict], path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "입찰참가이력"
    for row in to_rows(results):
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    rank_col = [label for _, label in COLUMNS].index("평가순위") + 1
    win_fill = PatternFill("solid", fgColor="FFF2CC")
    for row in ws.iter_rows(min_row=2):
        if str(row[rank_col - 1].value) == "1":
            for cell in row:
                cell.fill = win_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
